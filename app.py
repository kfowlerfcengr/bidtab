import os, io, json, re, tempfile
from datetime import datetime
from flask import Flask, request, jsonify, send_file, render_template_string

import anthropic
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.cell.cell import MergedCell

try:
    import pdfplumber
    HAS_PDF = True
except ImportError:
    HAS_PDF = False

app = Flask(__name__)
OUTPUT_DIR = tempfile.mkdtemp(prefix="bidtab_")

YELLOW = PatternFill("solid", fgColor="FFFF00")

NUMERIC_KEYWORDS = ["price","cost","qty","quantity","price_total","amount","adder","witnessed","test_cert","fee"]

def is_numeric_field(key):
    # Only match price_total exactly, not any field containing "total"
    if key == "price_total" or key.endswith("_price_total"):
        return True
    return any(kw in key for kw in NUMERIC_KEYWORDS if kw != "price_total")

def build_row_map(ws):
    """
    Build {row_number: field_key} by reading col A labels.
    Tracks current section to disambiguate duplicate labels.
    Handles both commercial sections (pricing/lead time) and
    multi-equipment templates (TLO/MLO/Hydraulic Start etc.)
    """
    row_map = {}

    # Section header labels — exact matches only, skip these rows
    SECTION_MARKERS = {
        "technical_commercial_bid_tab", "technical_bid_comparison",
        "basic_description", "construction",
        "special_requirements", "special_requirments",
        "motor_driver", "inspection_and_testing", "site_data",
        "other", "adders", "commercial_bid_comparison",
        "commercial_recommendationacceptability",
        "technical_recommendationacceptability",
        "commercial_recommendation_acceptability",
        "technical_recommendation_acceptability",
        "technical_recommendationacceptability",
        "recommendationacceptability",
        # Sub-section headers that appear inside equipment blocks
        "unit_data", "process_data", "performance_requirements",
        "design_criteria", "housing_mechanical_design", "scr_control",
        "bundle_details", "housing_materials", "electrical_data",
        "environment_data", "special_requirements", "physical_data",
        "nozzle_schedule",
        # Commercial section markers
        "pricing", "lead_time", "lead_time_weeks", "payment", "delivery", "misc", "materials",
    }

    # Equipment section markers — these name a piece of equipment, prefix subsequent rows
    # Key = normalized label, Value = short prefix to use
    EQUIPMENT_SECTIONS = {}  # filled dynamically from col A labels that look like equipment names

    # Commercial section prefixes
    SECTION_PREFIX_MAP = {
        "pricing":      "price",
        "pricing_":     "price",
        "lead_time":    "leadtime",
        "lead_time_":   "leadtime",
        "payment":      "payment",
        "delivery":     "delivery",
        "adders":       "adder",
    }

    current_section = ""       # commercial section context
    current_equipment = ""     # equipment block context (tlo, mlo, hydraulic_start, etc.)

    # First pass — find equipment section headers (col A labels that are NOT sub-sections
    # but appear as major dividers, i.e. rows followed by "UNIT DATA")
    rows_list = list(ws.iter_rows())
    equipment_header_rows = set()
    for i, row in enumerate(rows_list):
        label_cell = row[0] if row else None
        if not label_cell: continue
        label = str(label_cell.value or "").strip()
        if not label: continue
        key = re.sub(r'[^a-z0-9\s]', '', label.lower()).strip()
        key = re.sub(r'\s+', '_', key)
        # Look ahead — if next non-empty row is "UNIT DATA", this is an equipment header
        for j in range(i+1, min(i+4, len(rows_list))):
            next_cell = rows_list[j][0] if rows_list[j] else None
            if next_cell and next_cell.value:
                next_key = re.sub(r'[^a-z0-9\s]', '', str(next_cell.value).lower()).strip()
                next_key = re.sub(r'\s+', '_', next_key)
                if next_key == "unit_data":
                    equipment_header_rows.add(label_cell.row)
                    EQUIPMENT_SECTIONS[label_cell.row] = key[:12]  # short prefix
                break

    # Second pass — build the row map with full context awareness
    for row in ws.iter_rows():
        label_cell = row[0] if row else None
        if not label_cell or isinstance(label_cell, MergedCell):
            continue
        label = str(label_cell.value or "").strip()
        if not label or label.lower() in ("none", ""):
            continue

        key = label.lower()
        key = re.sub(r'[^a-z0-9\s]', '', key)
        key = re.sub(r'\s+', '_', key.strip())
        key = re.sub(r'^_+|_+$', '', key)
        key = re.sub(r'_+', '_', key)

        # Update equipment context
        if label_cell.row in equipment_header_rows:
            current_equipment = EQUIPMENT_SECTIONS[label_cell.row]
            continue  # don't map equipment header rows

        # Reset equipment context when we hit commercial sections
        if key in SECTION_PREFIX_MAP or key in {
            "commercial_bid_comparison", "commercial_recommendationacceptability",
            "technical_recommendationacceptability", "recommendationacceptability",
            "commercial_recommendation_acceptability", "technical_recommendation_acceptability",
        }:
            current_equipment = ""

        # Check if this is a section marker — update context but don't map
        if key in SECTION_MARKERS:
            for sec in SECTION_PREFIX_MAP:
                if key.startswith(sec) or key == sec:
                    current_section = sec
                    break
            continue

        # Apply equipment prefix for duplicate fields across equipment blocks
        if current_equipment:
            prefix = current_equipment
            if not key.startswith(prefix):
                key = f"{prefix}_{key}"

        # Apply commercial section prefix (overrides equipment prefix for commercial rows)
        if current_section in SECTION_PREFIX_MAP and not current_equipment:
            prefix = SECTION_PREFIX_MAP[current_section]
            if key != "quantity" and not key.startswith(prefix):
                key = f"{prefix}_{key}"

        if key:
            row_map[label_cell.row] = key

    return row_map

def build_system_prompt(row_map):
    field_list = ", ".join(sorted(set(row_map.values())))

    # Build a dynamic example from the actual row map instead of hardcoding pump fields
    price_fields = [f for f in row_map.values() if f.startswith("price_") and "total" not in f]
    leadtime_fields = [f for f in row_map.values() if f.startswith("leadtime_")]
    payment_fields = [f for f in row_map.values() if f.startswith("payment_")]
    delivery_fields = [f for f in row_map.values() if f.startswith("delivery_")]
    adder_fields = [f for f in row_map.values() if f.startswith("adder_")]

    field_hints = []
    if price_fields:
        field_hints.append(f"- Price fields ({', '.join(price_fields[:3])}) = equipment prices as plain numbers, no $ or commas")
    if leadtime_fields:
        field_hints.append(f"- Lead time fields ({', '.join(leadtime_fields[:2])}) = delivery time (e.g. '12-14 Weeks ARO')")
    if payment_fields:
        field_hints.append(f"- Payment fields ({', '.join(payment_fields[:2])}) = payment terms and schedule")
    if delivery_fields:
        field_hints.append(f"- Delivery fields ({', '.join(delivery_fields[:3])}) = shipping terms, location, warranty, comments")
    if adder_fields:
        field_hints.append(f"- Adder fields ({', '.join(adder_fields[:3])}) = optional add-on prices as plain numbers")
    if "quantity" in row_map.values():
        field_hints.append("- 'quantity' = number of units (integer)")

    field_hints_str = "\n".join(field_hints) if field_hints else ""

    # Detect if this is a multi-equipment template
    equip_prefixes = sorted(set(
        f.split('_')[0] + '_' + f.split('_')[1] if len(f.split('_')) > 1 else f.split('_')[0]
        for f in row_map.values()
        if '_' in f and not any(f.startswith(p) for p in ('price_','leadtime_','payment_','delivery_','adder_'))
    ))
    multi_equip_note = ""
    if len(equip_prefixes) > 1:
        multi_equip_note = f"""
IMPORTANT — this template covers MULTIPLE equipment types with separate sections.
Each equipment section has its own prefixed fields (e.g. {', '.join(equip_prefixes[:4])}).
Map each vendor's data for each piece of equipment to the CORRECT prefixed field.
Do NOT put data from one equipment type into another equipment type's fields.
Each vendor may quote different equipment — if a vendor did not quote a specific item, use null for that item's fields."""

    return f"""You are a THOROUGH technical bid tab assistant for an engineering procurement team.

You may receive ONE OR MORE DATA SHEETS for engineering equipment, and up to SEVERAL VENDOR QUOTES.
The equipment type will vary — pumps, heaters, compressors, heat exchangers, vessels, or any other equipment.
Extract all available data and return ONLY valid JSON — no markdown, no explanation.

Return this exact structure:
{{
  "datasheet": {{ ...combined fields from ALL data sheets... }},
  "vendor1": {{ ...fields from vendor 1... }},
  "vendor2": {{ ...fields from vendor 2... }},
  "vendor3": {{ ...fields from vendor 3... }}
}}

If there are multiple DATA SHEETS, merge all their fields into the single "datasheet" object.
If the same field appears in multiple data sheets, use the most specific or detailed value.

The template has these exact field keys — map ALL extracted data to these key names:
{field_list}

How the field keys are structured (all come from this specific template):
{field_hints_str}
{multi_equip_note}

General rules for field naming:
- Fields prefixed with "price_" are equipment/component prices → numbers only (no $ or commas)
- Fields prefixed with "leadtime_" are delivery lead times → text (e.g. "12-14 Weeks ARO")
- Fields prefixed with "payment_" are payment terms and schedules → text
- Fields prefixed with "delivery_" are shipping, manufacturing, warranty, and commercial info → text
- Fields prefixed with "adder_" are optional add-on prices → numbers only
- "quantity" = number of units → integer

Extraction rules:
- "datasheet" key = combined values from ALL DATA SHEETS only
- "vendor1/2/3/n" = values from each vendor quote only — never mix sources
- price_* and adder_* fields: plain numbers only (no $, commas, or units)
- quantity field: integer only
- ALL OTHER fields: preserve units exactly as they appear in the source (e.g. "3 kW", "43 lbs", "26.5 in", "122°F (50°C)", "35 business days")
- For dimension fields: preserve full decimal precision — never round (26.5 stays 26.5, not 26)
- For temperature fields: include both °F and °C if the source has both
- For electrical capacity: include unit (e.g. "3 kW" not just "3")
- For weight fields: include unit (e.g. "43 lbs" not just "43")
- For voltage/phase/frequency: extract the FULL string including Hz (e.g. "480V/3PH/60HZ" — never truncate to "480V/3PH")
- For hazardous area class: extract the area classification exactly as written by each vendor — do not confuse certification names (CSA, ATEX) with area class (e.g. "CLASS I DIV 1 GR. B,C,D" or "ZONE 2, GROUP IIA,B")
- For design_pressure and design_tempurature fields: ONLY put pressure/temperature values here (e.g. "15 PSI", "78°C") — NEVER put certification names like "CSA" or "ATEX" in these fields
- For price fields: if a vendor quotes BOTH a main equipment unit AND a control panel/PCA as separate line items for the SAME piece of equipment, ADD them together for that equipment's price field
- For lead times: use the PRODUCT DELIVERY lead time (not drawing submittal). If one lead time covers all items, repeat it for each item field
- For "total_length_dimension" fields: use the full immersion/bundle length (the long dimension)
- For "flange_face_to_bundle_end_dim" fields: use the shorter cold section / unheated length
- For design pressure/temperature: extract from vendor quotes when stated
- quotation_number = the vendor's quote/reference number, NOT their internal project reference code
- null for any field not found — do NOT fabricate or guess
- Extract every spec, technical parameter, price, commercial term, and note you can find
- Return ONLY the JSON object, nothing else"""

def extract_text(file_storage) -> str:
    name = (file_storage.filename or "").lower()
    data = file_storage.read()

    if name.endswith(".pdf"):
        if not HAS_PDF:
            return "[PDF unavailable — install pdfplumber]"
        try:
            with pdfplumber.open(io.BytesIO(data)) as pdf:
                pages = []
                for i, page in enumerate(pdf.pages, 1):
                    text = page.extract_text()
                    if text and text.strip():
                        pages.append(f"[Page {i}]\n{text.strip()}")
                return "\n\n".join(pages) if pages else "[PDF appears to be empty or image-only]"
        except Exception as e:
            return f"[PDF read error: {e}]"

    if name.endswith((".xlsx", ".xls")):
        try:
            wb = load_workbook(io.BytesIO(data), data_only=True)
            lines = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                sheet_lines = []
                for row in ws.iter_rows(values_only=True):
                    parts = [str(c).strip() if c is not None else "" for c in row]
                    if any(p for p in parts):
                        sheet_lines.append("\t".join(parts))
                if sheet_lines:
                    lines.append(f"[Sheet: {sheet_name}]")
                    lines.extend(sheet_lines)
            return "\n".join(lines) if lines else "[Excel file appears empty]"
        except Exception as e:
            return f"[Excel read error: {e}]"

    if name.endswith(".csv"):
        try:
            return data.decode("utf-8", errors="replace")
        except Exception as e:
            return f"[CSV read error: {e}]"

    # Plain text fallback
    try:
        return data.decode("utf-8", errors="replace")
    except Exception as e:
        return f"[File read error: {e}]"

def coerce(val, field):
    if val is None: return None
    s = str(val).strip()
    if not s or s.lower() == "null": return None
    if is_numeric_field(field):
        # For price/adder/quantity fields: strip units, return plain number
        try:
            f = float(s.replace(",","").replace("$","").replace(" ",""))
            return int(f) if f == int(f) else f
        except:
            # If it can't be parsed as a number, return as-is (e.g. "TBD")
            return s
    # For all other fields: return the value as-is, preserving units
    return s

def fill_excel(template_bytes, data, pi):
    wb = load_workbook(io.BytesIO(template_bytes))
    ws = wb.active

    row_map = build_row_map(ws)

    # Safety check — if template has vendor data already filled in,
    # warn but continue (user may have uploaded a pre-filled template)
    vendor_filled_count = sum(
        1 for row in ws.iter_rows(min_row=8, max_row=ws.max_row)
        for cell in row
        if cell.column in (3, 4, 5)  # cols C, D, E
        and cell.value is not None
        and str(cell.value).strip() not in ("", "-", "By Vendor", "N/A", "Vendor 1", "Vendor 2", "Vendor 3")
        and not isinstance(cell, MergedCell)
    )
    # If template already has lots of vendor data, it's probably not blank
    # We still proceed but the row_map will be built from whatever is there

    def safe_write(addr, val):
        if not val: return
        cell = ws[addr]
        if not isinstance(cell, MergedCell):
            cell.value = val

    # Project header — hardcoded, same position in all F&C templates
    safe_write("C2", pi.get("client",""))
    safe_write("C3", pi.get("project_no",""))
    safe_write("C4", pi.get("project_name",""))
    safe_write("E4", datetime.today().strftime("%m/%d/%Y"))
    safe_write("C5", pi.get("location",""))
    safe_write("F5", pi.get("author",""))
    safe_write("B7", pi.get("equipment",""))
    safe_write("C7", pi.get("vendor1_name","Vendor 1"))
    safe_write("D7", pi.get("vendor2_name","Vendor 2"))
    safe_write("E7", pi.get("vendor3_name","Vendor 3"))

    # Data rows — dynamic, works for any template
    col_map = {"datasheet":"B","vendor1":"C","vendor2":"D","vendor3":"E"}
    for row_num, field in row_map.items():
        if row_num <= 7:
            continue  # header rows handled above, don't overwrite
        for src, col in col_map.items():
            val = coerce((data.get(src) or {}).get(field), field)
            cell = ws[f"{col}{row_num}"]
            if isinstance(cell, MergedCell):
                continue
            if val is not None:
                cell.value = val
            elif col in ("C","D","E"):
                curr = cell.value
                if curr is None or str(curr).strip() in ("","-","By Vendor","N/A","by vendor","n/a"):
                    cell.fill = YELLOW

    # TOTAL formula — sum only the price rows (above total, above quantity row)
    # Find price rows, quantity row, and total row precisely
    price_rows = [r for r, f in row_map.items() if f.startswith("price_") and "total" not in f]
    qty_rows = [r for r, f in row_map.items() if f == "quantity"]
    total_rows = [r for r, f in row_map.items() if "price_total" in f or f == "total"]

    for row_num in total_rows:
        qty_row = qty_rows[0] if qty_rows else None
        if price_rows and qty_row:
            min_price = min(price_rows)
            max_price = max(price_rows)
            for col in ("C","D","E"):
                cell = ws[f"{col}{row_num}"]
                if not isinstance(cell, MergedCell):
                    cell.value = f"=SUM({col}{min_price}:{col}{max_price})*{col}{qty_row}"

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read(), row_map


HTML = r"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Bid Tab Agent</title>
<style>
@import url('https://fonts.googleapis.com/css2?family=IBM+Plex+Mono:wght@400;500&family=IBM+Plex+Sans:wght@300;400;500;600&display=swap');
:root{
  --ink:#0f1117;--ink2:#3a3d4a;--ink3:#6b7080;
  --line:#e2e4ea;--line2:#c8cad4;--bg:#f7f8fa;--bg2:#fff;
  --accent:#e8286a;--accent2:#fdedf4;
  --accent-warm:#f5a623;
  --green:#0d6e3b;--gbg:#e6f4ed;
  --red:#9b1c1c;--rbg:#fde8e8;
  --mono:'IBM Plex Mono',monospace;--sans:'IBM Plex Sans',sans-serif;
}
*,*::before,*::after{box-sizing:border-box;margin:0;padding:0;}
body{font-family:var(--sans);background:var(--bg);color:var(--ink);font-size:14px;line-height:1.5;}
.topbar{background:var(--ink);color:#fff;height:58px;padding:0 28px;display:flex;align-items:center;justify-content:space-between;position:sticky;top:0;z-index:100;border-bottom:1px solid rgba(255,255,255,.06);}
.brand{display:flex;align-items:center;gap:10px;}
.brand img{height:34px;width:auto;display:block;mix-blend-mode:screen;}
.brand-name{font-family:var(--sans);font-size:15px;font-weight:500;color:#fff;letter-spacing:.01em;}
.brand-name span{color:rgba(255,255,255,.38);font-weight:300;font-size:12px;margin-left:6px;letter-spacing:.04em;}
.badge{font-family:var(--mono);font-size:11px;padding:4px 12px;border-radius:20px;background:rgba(255,255,255,.08);color:rgba(255,255,255,.5);}
.page{max-width:780px;margin:0 auto;padding:36px 20px 80px;}
.card{background:var(--bg2);border:1px solid var(--line);border-radius:12px;margin-bottom:18px;overflow:hidden;}
.ch{padding:15px 22px;border-bottom:1px solid var(--line);display:flex;align-items:center;gap:12px;}
.ch-ico{width:28px;height:28px;border-radius:50%;background:linear-gradient(135deg,var(--accent-warm),var(--accent));color:#fff;font-family:var(--mono);font-size:11px;display:flex;align-items:center;justify-content:center;flex-shrink:0;font-weight:500;}
.ch h2{font-size:14px;font-weight:600;margin-bottom:2px;}
.ch p{font-size:11px;color:var(--ink3);}
.cb{padding:18px 22px;}
.igrid{display:grid;grid-template-columns:1fr 1fr;gap:13px;}
.fld label{display:block;font-size:10px;font-family:var(--mono);color:var(--ink3);text-transform:uppercase;letter-spacing:.06em;margin-bottom:4px;}
.fld input{width:100%;border:1px solid var(--line2);border-radius:7px;padding:8px 11px;font-family:var(--sans);font-size:13px;color:var(--ink);outline:none;}
.fld input:focus{border-color:var(--accent);}
.uz{border:2px dashed var(--line2);border-radius:10px;padding:24px 18px;text-align:center;cursor:pointer;transition:all .2s;background:var(--bg);}
.uz:hover,.uz.drag{border-color:var(--accent);background:var(--accent2);}
.uz input{display:none;}
.uz .uico{font-size:24px;margin-bottom:6px;}
.uz h4{font-size:13px;font-weight:500;margin-bottom:2px;}
.uz p{font-size:11px;color:var(--ink3);}
.fchip{display:inline-flex;align-items:center;gap:6px;background:var(--gbg);border:1px solid #a3d9b8;color:var(--green);border-radius:6px;padding:4px 10px;font-family:var(--mono);font-size:11px;margin-top:6px;}
.vgrid{display:grid;grid-template-columns:repeat(auto-fill,minmax(190px,1fr));gap:12px;}
.vcard{border:1px solid var(--line2);border-radius:10px;overflow:hidden;}
.vcard.loaded{border-color:#f5a19d;}
.vhead{background:var(--bg);padding:8px 12px;border-bottom:1px solid var(--line);font-family:var(--mono);font-size:10px;color:var(--ink3);text-transform:uppercase;letter-spacing:.05em;display:flex;align-items:center;justify-content:space-between;}
.vhead.loaded{background:#fff5f8;color:var(--accent);border-bottom-color:#f5a19d;}
.vremove{background:none;border:none;color:var(--ink3);cursor:pointer;font-size:15px;line-height:1;padding:0 2px;border-radius:3px;font-family:var(--sans);}
.vremove:hover{color:var(--red);background:var(--rbg);}
.vbody{padding:11px 12px;background:var(--bg2);}
.vbody input{width:100%;border:1px solid var(--line2);border-radius:6px;padding:7px 10px;font-size:13px;font-family:var(--sans);color:var(--ink);outline:none;margin-bottom:8px;}
.vbody input:focus{border-color:var(--accent);}
.vup{width:100%;padding:7px;border:1px dashed var(--line2);border-radius:6px;background:var(--bg);font-size:11px;color:var(--ink3);cursor:pointer;text-align:center;font-family:var(--sans);transition:all .15s;}
.vup:hover{border-color:var(--accent);color:var(--accent);background:var(--accent2);}
.vfname{font-family:var(--mono);font-size:10px;color:var(--green);margin-top:5px;min-height:13px;}
.add-vbtn{display:block;width:100%;margin-top:10px;padding:8px;background:var(--bg2);border:1px dashed var(--line2);border-radius:8px;color:var(--ink3);font-size:12px;cursor:pointer;font-family:var(--sans);transition:all .15s;}
.add-vbtn:hover{border-color:var(--accent);color:var(--accent);background:var(--accent2);}
.sub-area{margin-top:24px;text-align:center;}
.sub-btn{background:linear-gradient(135deg,var(--accent-warm),var(--accent));color:#fff;border:none;padding:13px 42px;border-radius:8px;font-size:15px;font-weight:500;font-family:var(--sans);cursor:pointer;transition:opacity .15s;}
.sub-btn:hover{opacity:.88;}
.sub-btn:disabled{background:linear-gradient(135deg,#ccc,#aaa);cursor:not-allowed;opacity:1;}
.sub-note{font-size:11px;color:var(--ink3);margin-top:8px;}
.rbox{background:var(--gbg);border:1px solid #a3d9b8;border-radius:10px;padding:22px;margin-top:22px;display:none;text-align:center;}
.rbox.show{display:block;}
.rbox h3{font-size:15px;font-weight:600;color:var(--green);margin-bottom:6px;}
.rbox p{font-size:12px;color:#1a5c35;margin-bottom:14px;}
.dl-btn{display:inline-flex;align-items:center;gap:7px;background:var(--green);color:#fff;padding:10px 26px;border-radius:8px;text-decoration:none;font-size:13px;font-weight:500;}
.dl-btn:hover{background:#0a5a30;}
.ebox{background:var(--rbg);border:1px solid #f5c2c2;border-radius:10px;padding:14px 18px;margin-top:18px;display:none;}
.ebox.show{display:block;}
.ebox p{color:var(--red);font-size:13px;}
.fbox{background:#fff;border:1px solid var(--line);border-radius:12px;margin-top:18px;display:none;overflow:hidden;}
.fbox.show{display:block;}
.fbox-head{background:var(--bg);padding:14px 22px;border-bottom:1px solid var(--line);display:flex;align-items:baseline;gap:10px;}
.fbox-head span:first-child{font-weight:600;font-size:14px;}
.fbox-sub{font-size:11px;color:var(--ink3);}
.fbox-body{padding:18px 22px;}
.rating-row{display:flex;gap:10px;margin-bottom:16px;}
.rating-btn{flex:1;padding:10px 8px;border:2px solid var(--line2);border-radius:8px;background:var(--bg);font-size:13px;cursor:pointer;font-family:var(--sans);transition:all .15s;font-weight:500;}
.rating-btn:hover{border-color:var(--ink3);background:#f0f0f0;}
.rating-btn.active-bad{border-color:#e53e3e;background:#fff5f5;color:#c53030;}
.rating-btn.active-track{border-color:#d69e2e;background:#fffff0;color:#b7791f;}
.rating-btn.active-good{border-color:#38a169;background:#f0fff4;color:#276749;}
.fbox-fields{display:flex;flex-direction:column;gap:12px;margin-bottom:16px;}
.fbox-fields .fld label{display:block;font-size:10px;font-family:var(--mono);color:var(--ink3);text-transform:uppercase;letter-spacing:.06em;margin-bottom:4px;}
.fbox-fields textarea{width:100%;border:1px solid var(--line2);border-radius:7px;padding:8px 11px;font-family:var(--sans);font-size:13px;color:var(--ink);outline:none;resize:vertical;}
.fbox-fields textarea:focus{border-color:var(--accent);}
.fbox-actions{display:flex;align-items:center;gap:12px;}
.fb-submit{background:var(--accent);color:#fff;border:none;padding:10px 28px;border-radius:8px;font-size:14px;font-weight:500;font-family:var(--sans);cursor:pointer;transition:opacity .15s;}
.fb-submit:hover{opacity:.88;}
.fb-submit:disabled{background:#aaa;cursor:not-allowed;opacity:1;}
.fb-skip{background:none;border:none;color:var(--ink3);font-size:13px;cursor:pointer;font-family:var(--sans);text-decoration:underline;}
.fb-thanks{display:none;color:var(--green);font-size:13px;font-weight:500;margin-top:10px;}
.overlay{position:fixed;inset:0;background:rgba(15,17,23,.72);display:none;align-items:center;justify-content:center;z-index:200;}
.overlay.show{display:flex;}
.pbox{background:#fff;border-radius:14px;padding:38px 46px;text-align:center;max-width:360px;width:90%;}
.spin{width:42px;height:42px;border:3px solid var(--line);border-top-color:var(--accent);border-radius:50%;animation:spin .7s linear infinite;margin:0 auto 18px;}
@keyframes spin{to{transform:rotate(360deg);}}
.pbox h3{font-size:15px;font-weight:600;margin-bottom:5px;}
.pbox p{font-size:12px;color:var(--ink3);}
.pstep{font-family:var(--mono);font-size:11px;color:var(--accent);margin-top:10px;}
</style>
</head>
<body>
<div class="topbar">
  <div class="brand">
    <img src="/static/logo.png?v=3" alt="F&C Engineers">
    <span class="brand-name">Bid Tab Agent<span>AI</span></span>
  </div>
  <div class="badge" id="badge">Ready</div>
</div>
<div class="page">

  <div class="card">
    <div class="ch"><div class="ch-ico">📁</div><div><h2>Project Info</h2><p>Fills the header of your bid tab</p></div></div>
    <div class="cb">
      <div class="igrid">
        <div class="fld"><label>Client</label><input id="pi-client" placeholder="e.g. FTAI"></div>
        <div class="fld"><label>Project No.</label><input id="pi-projno" placeholder="e.g. FTA260107"></div>
        <div class="fld"><label>Project Name</label><input id="pi-name" placeholder="e.g. FTAI CFM56 Package"></div>
        <div class="fld"><label>Project Location</label><input id="pi-loc" placeholder="e.g. Mobile"></div>
        <div class="fld"><label>Equipment / Item</label><input id="pi-equip" placeholder="e.g. Pumps"></div>
        <div class="fld"><label>Author</label><input id="pi-author" placeholder="Your name"></div>
      </div>
    </div>
  </div>

  <div class="card">
    <div class="ch"><div class="ch-ico">1</div><div><h2>Bid Tab Template</h2><p>Your blank .xlsx — formatting, colors and merges will be preserved exactly</p></div></div>
    <div class="cb">
      <div class="uz" id="z-tmpl" onclick="document.getElementById('f-tmpl').click()"
           ondragover="zo(event,'z-tmpl')" ondragleave="zl('z-tmpl')" ondrop="zd(event,'z-tmpl','f-tmpl')">
        <input type="file" id="f-tmpl" accept=".xlsx" onchange="chip(this,'chip-tmpl')">
        <div class="uico">📋</div><h4>Upload your blank bid tab template (.xlsx)</h4>
        <p>Required — all original formatting will be preserved</p>
        <div id="chip-tmpl"></div>
      </div>
    </div>
  </div>

  <div class="card">
    <div class="ch"><div class="ch-ico">2</div><div><h2>Data Sheets</h2><p>Engineering specs — fills the standards column. Add one per equipment type.</p></div></div>
    <div class="cb">
      <div class="vgrid" id="dsgrid"></div>
      <button type="button" class="add-vbtn" onclick="addDataSheet()">+ Add Data Sheet</button>
    </div>
  </div>

  <div class="card">
    <div class="ch"><div class="ch-ico">3</div><div><h2>Vendor Quotes</h2><p>Any format, any layout — AI figures it out</p></div></div>
    <div class="cb">
      <div class="vgrid" id="vgrid"></div>
      <button type="button" class="add-vbtn" onclick="addVendor()">+ Add Vendor</button>
    </div>
  </div>

  <div class="sub-area">
    <button class="sub-btn" id="sub-btn" onclick="run()">Generate Bid Tab</button>
    <div class="sub-note">Server reads your files with AI → returns a perfectly formatted .xlsx</div>
  </div>

  <div class="rbox" id="rbox">
    <h3>✓ Bid tab ready</h3>
    <p id="rmsg"></p>
    <a class="dl-btn" id="dl-link" href="#" onclick="showFeedback()">⬇ Download Filled Bid Tab (.xlsx)</a>
  </div>
  <div class="ebox" id="ebox"><p id="emsg"></p></div>

  <!-- Feedback panel — appears after download -->
  <div class="fbox" id="fbox">
    <div class="fbox-head">
      <span>📋 How did it do?</span>
      <span class="fbox-sub">Your feedback teaches the agent to improve on this template</span>
    </div>
    <div class="fbox-body">
      <div class="rating-row">
        <button class="rating-btn" id="rb-bad"    onclick="setRating('bad')">🔴 Bad</button>
        <button class="rating-btn" id="rb-track"  onclick="setRating('track')">🟡 On the right track</button>
        <button class="rating-btn" id="rb-good"   onclick="setRating('good')">🟢 Good</button>
      </div>
      <div class="fbox-fields" id="fbox-fields">
        <div class="fld">
          <label>What was wrong / what was good?</label>
          <textarea id="fb-comment" placeholder="e.g. MLO weight was 43 lbs but should be 77 lbs. Lead times were correct." rows="3"></textarea>
        </div>
        <div class="fld">
          <label>Specific cell corrections (optional)</label>
          <textarea id="fb-cells" placeholder="e.g. C173 = 77 lbs&#10;C278 = 11515&#10;D152 = 8 kW" rows="3"></textarea>
        </div>
      </div>
      <div class="fbox-actions">
        <button class="fb-submit" id="fb-submit" onclick="submitFeedback()" disabled>Submit Feedback</button>
        <button class="fb-skip" onclick="skipFeedback()">Skip</button>
      </div>
      <div class="fb-thanks" id="fb-thanks">✓ Feedback saved — the agent will use this next time.</div>
    </div>
  </div>
</div>

<div class="overlay" id="overlay">
  <div class="pbox">
    <div class="spin"></div>
    <h3>Working on it…</h3>
    <p id="pmsg">Reading your files</p>
    <div class="pstep" id="pstep"></div>
  </div>
</div>

<script>
/* ── Feedback system ── */
let currentRating=null,currentFilename=null,currentEquipment=null;

function showFeedback(){
  const href=document.getElementById('dl-link').getAttribute('data-href');
  if(href){const a=document.createElement('a');a.href=href;a.download=document.getElementById('dl-link').getAttribute('data-download');a.click();}
  setTimeout(()=>{const fb=document.getElementById('fbox');fb.classList.add('show');fb.scrollIntoView({behavior:'smooth',block:'nearest'});},400);
  return false;
}
function setRating(r){
  currentRating=r;
  ['bad','track','good'].forEach(x=>document.getElementById('rb-'+x).className='rating-btn'+(r===x?' active-'+x:''));
  document.getElementById('fb-submit').disabled=false;
}
async function submitFeedback(){
  if(!currentRating)return;
  const comment=document.getElementById('fb-comment').value.trim();
  const cells=document.getElementById('fb-cells').value.trim();
  document.getElementById('fb-submit').disabled=true;
  try{
    await fetch('/feedback',{method:'POST',headers:{'Content-Type':'application/json'},
      body:JSON.stringify({rating:currentRating,comment,cell_corrections:cells,filename:currentFilename,equipment:currentEquipment,timestamp:new Date().toISOString()})});
    const thanks=document.getElementById('fb-thanks');
    const fields=document.getElementById('fbox-fields');
    const actions=document.getElementById('fbox-actions');
    if(thanks)thanks.style.display='block';
    if(fields)fields.style.display='none';
    if(actions)actions.style.display='none';
  }catch(e){document.getElementById('fb-submit').disabled=false;alert('Could not save feedback: '+e.message);}
}
function skipFeedback(){document.getElementById('fbox').classList.remove('show');}

function zo(e,id){e.preventDefault();document.getElementById(id).classList.add('drag');}
function zl(id){document.getElementById(id).classList.remove('drag');}
function zd(e,zid,fid){e.preventDefault();zl(zid);const f=e.dataTransfer.files[0];if(!f)return;const inp=document.getElementById(fid);const dt=new DataTransfer();dt.items.add(f);inp.files=dt.files;chip(inp,fid==='f-ds'?'chip-ds':fid==='f-tmpl'?'chip-tmpl':null);}
function chip(inp,chipId){const f=inp.files[0];if(!f||!chipId)return;document.getElementById(chipId).innerHTML=`<div class="fchip">✓ ${f.name}</div>`;}

let dsSlots=[];
let nextDSIdx=0;

function makeDSCard(idx, displayNum){
  const div=document.createElement('div');
  div.className='vcard';div.id='dsc'+idx;
  div.innerHTML=`
    <div class="vhead" id="dsh${idx}">
      <span>Data Sheet ${displayNum}</span>
      <button type="button" class="vremove" onclick="removeDS(${idx})" title="Remove">×</button>
    </div>
    <div class="vbody">
      <div class="vup" onclick="document.getElementById('dsf${idx}').click()">+ Attach file</div>
      <input type="file" id="dsf${idx}" style="display:none" accept=".xlsx,.xls,.csv,.txt,.pdf"
             onchange="dsfile(this,${idx})">
      <div class="vfname" id="dsfn${idx}"></div>
    </div>`;
  return div;
}

function addDataSheet(){
  const idx=nextDSIdx++;
  dsSlots.push(idx);
  document.getElementById('dsgrid').appendChild(makeDSCard(idx,dsSlots.length));
  syncDSButtons();
}

function removeDS(idx){
  if(dsSlots.length<=1)return;
  dsSlots=dsSlots.filter(i=>i!==idx);
  const card=document.getElementById('dsc'+idx);
  if(card)card.remove();
  syncDSButtons();
}

function syncDSButtons(){
  const show=dsSlots.length>1;
  dsSlots.forEach(i=>{
    const btn=document.getElementById('dsc'+i)?.querySelector('.vremove');
    if(btn)btn.style.visibility=show?'visible':'hidden';
  });
}

function dsfile(inp,idx){
  const f=inp.files[0];if(!f)return;
  document.getElementById('dsfn'+idx).textContent='✓ '+f.name;
  document.getElementById('dsc'+idx).classList.add('loaded');
  document.getElementById('dsh'+idx).classList.add('loaded');
}

(function initDS(){addDataSheet();}());

let vendorSlots=[];
let nextVIdx=0;

function makeVendorCard(idx,displayNum){
  const div=document.createElement('div');
  div.className='vcard';div.id='vc'+idx;
  div.innerHTML=`
    <div class="vhead" id="vh${idx}">
      <span id="vhl${idx}">Vendor ${displayNum}</span>
      <button type="button" class="vremove" id="vrm${idx}" onclick="removeVendor(${idx})" title="Remove vendor">×</button>
    </div>
    <div class="vbody">
      <input id="vn${idx}" placeholder="Vendor name" value="Vendor ${displayNum}"
             oninput="document.getElementById('vhl${idx}').textContent=this.value||'Vendor ${displayNum}'">
      <div class="vup" onclick="document.getElementById('vf${idx}').click()">+ Attach quote</div>
      <input type="file" id="vf${idx}" style="display:none" accept=".xlsx,.xls,.csv,.txt,.pdf"
             onchange="vfile(this,${idx})">
      <div class="vfname" id="vfn${idx}"></div>
    </div>`;
  return div;
}

function addVendor(){
  const idx=nextVIdx++;
  vendorSlots.push(idx);
  document.getElementById('vgrid').appendChild(makeVendorCard(idx,vendorSlots.length));
  syncRemoveButtons();
}

function removeVendor(idx){
  if(vendorSlots.length<=1)return;
  vendorSlots=vendorSlots.filter(i=>i!==idx);
  const card=document.getElementById('vc'+idx);
  if(card)card.remove();
  syncRemoveButtons();
}

function syncRemoveButtons(){
  const show=vendorSlots.length>1;
  vendorSlots.forEach(i=>{
    const btn=document.getElementById('vrm'+i);
    if(btn)btn.style.visibility=show?'visible':'hidden';
  });
}

function vfile(inp,idx){
  const f=inp.files[0];if(!f)return;
  document.getElementById('vfn'+idx).textContent='✓ '+f.name;
  document.getElementById('vc'+idx).classList.add('loaded');
  document.getElementById('vh'+idx).classList.add('loaded');
}

(function initVendors(){for(let i=0;i<3;i++)addVendor();}());

async function run(){
  const tmplFile=document.getElementById('f-tmpl').files[0];
  if(!tmplFile){alert('Please upload your bid tab template (.xlsx).');return;}
  const hasVendor=vendorSlots.some(i=>document.getElementById('vf'+i)?.files[0]);
  const hasDS=dsSlots.some(i=>document.getElementById('dsf'+i)?.files[0]);
  if(!hasDS&&!hasVendor){alert('Please upload at least a data sheet or one vendor quote.');return;}

  document.getElementById('sub-btn').disabled=true;
  document.getElementById('rbox').classList.remove('show');
  document.getElementById('ebox').classList.remove('show');
  document.getElementById('overlay').classList.add('show');
  document.getElementById('badge').textContent='Processing…';

  const msgs=['Reading files…','AI extracting specs…','Mapping vendor data…','Writing Excel…'];
  let mi=0;
  const ticker=setInterval(()=>{
    document.getElementById('pmsg').textContent=msgs[Math.min(mi,msgs.length-1)];
    document.getElementById('pstep').textContent=`Step ${Math.min(mi+1,4)} of 4`;
    mi++;
  },5000);

  try{
    const fd=new FormData();
    fd.append('template',tmplFile);

    // Data sheets
    dsSlots.forEach((slotIdx,pos)=>{
      const file=document.getElementById('dsf'+slotIdx)?.files[0];
      if(file) fd.append(pos===0?'datasheet':`datasheet_${pos+1}`,file);
    });
    fd.append('datasheet_count',dsSlots.length);

    vendorSlots.forEach((slotIdx,pos)=>{
      const name=document.getElementById('vn'+slotIdx)?.value||`Vendor ${pos+1}`;
      const file=document.getElementById('vf'+slotIdx)?.files[0];
      fd.append(`vendor_name_${pos+1}`,name);
      if(file)fd.append(`vendor_${pos+1}`,file);
    });
    fd.append('vendor_count',vendorSlots.length);

    fd.append('client',document.getElementById('pi-client').value);
    fd.append('project_no',document.getElementById('pi-projno').value);
    fd.append('project_name',document.getElementById('pi-name').value);
    fd.append('location',document.getElementById('pi-loc').value);
    fd.append('equipment',document.getElementById('pi-equip').value);
    fd.append('author',document.getElementById('pi-author').value);

    const resp=await fetch('/generate',{method:'POST',body:fd});
    const result=await resp.json();
    clearInterval(ticker);
    document.getElementById('overlay').classList.remove('show');
    document.getElementById('sub-btn').disabled=false;
    document.getElementById('badge').textContent='Ready';

    if(result.error){
      document.getElementById('emsg').textContent='Error: '+result.error;
      document.getElementById('ebox').classList.add('show');
    }else{
      const versionLabel = result.version ? ` — Version ${result.version}` : '';
      document.getElementById('rmsg').textContent=result.message+versionLabel;
      // Store for feedback
      currentFilename=result.filename;
      currentEquipment=document.getElementById('pi-equip').value||'Unknown';
      // Wire download link
      const dlLink=document.getElementById('dl-link');
      dlLink.setAttribute('data-href','/download/'+result.filename);
      dlLink.setAttribute('data-download',result.filename);
      dlLink.href='#';
      dlLink.onclick=showFeedback;
      document.getElementById('rbox').classList.add('show');
      // Reset feedback panel state
      currentRating=null;
      ['bad','track','good'].forEach(x=>document.getElementById('rb-'+x).className='rating-btn');
      document.getElementById('fb-submit').disabled=true;
      document.getElementById('fb-comment').value='';
      document.getElementById('fb-cells').value='';
      const fbThanks=document.getElementById('fb-thanks');
      const fbFields=document.getElementById('fbox-fields');
      const fbActions=document.getElementById('fbox-actions');
      if(fbThanks)fbThanks.style.display='none';
      if(fbFields)fbFields.style.display='';
      if(fbActions)fbActions.style.display='';
      document.getElementById('fbox').classList.remove('show');
    }
  }catch(err){
    clearInterval(ticker);
    document.getElementById('overlay').classList.remove('show');
    document.getElementById('sub-btn').disabled=false;
    document.getElementById('badge').textContent='Error';
    document.getElementById('emsg').textContent='Error: '+err.message;
    document.getElementById('ebox').classList.add('show');
  }
}
</script>
</body>
</html>"""


FEEDBACK_DIR = os.path.join(tempfile.gettempdir(), "bidtab_feedback")
os.makedirs(FEEDBACK_DIR, exist_ok=True)

# Seed feedback store from bundled seed files on first run
def _seed_feedback():
    seed_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".")
    for fname in os.listdir(seed_dir):
        if not fname.endswith("_feedback_seed.json"):
            continue
        equip_slug = fname.replace("_feedback_seed.json", "")
        dest = os.path.join(FEEDBACK_DIR, f"{equip_slug}.json")
        if os.path.exists(dest):
            continue  # don't overwrite existing feedback
        try:
            import shutil
            shutil.copy(os.path.join(seed_dir, fname), dest)
        except Exception:
            pass
_seed_feedback()

def load_feedback_for_template(equipment: str) -> str:
    """Load all past feedback for this equipment type and format it for Claude."""
    slug = re.sub(r'[^a-z0-9]', '_', equipment.lower()).strip('_')
    fpath = os.path.join(FEEDBACK_DIR, f"{slug}.json")
    if not os.path.exists(fpath):
        return ""
    try:
        with open(fpath) as f:
            entries = json.load(f)
    except Exception:
        return ""
    if not entries:
        return ""

    # Build prompt injection from feedback history
    lines = [f"\nPAST FEEDBACK ON THIS TEMPLATE ({equipment}) — {len(entries)} sessions:"]
    # Weight recent feedback more — show last 10, summarise older
    recent = entries[-10:]
    for e in recent:
        rating_label = {"bad": "❌ BAD", "track": "⚠️ ON THE RIGHT TRACK", "good": "✅ GOOD"}.get(e.get("rating",""), "?")
        lines.append(f"\n  [{rating_label}] {e.get('timestamp','')[:10]}")
        if e.get("comment"):
            lines.append(f"    Feedback: {e['comment']}")
        if e.get("cell_corrections"):
            lines.append(f"    Cell corrections: {e['cell_corrections']}")

    # Extract all cell corrections as hard rules
    all_corrections = []
    for e in entries:
        if e.get("cell_corrections"):
            all_corrections.append(e["cell_corrections"])
    if all_corrections:
        lines.append(f"\n  KNOWN CORRECTIONS (apply these every time):")
        for c in all_corrections[-5:]:  # last 5 correction sets
            for line in c.strip().split('\n'):
                if '=' in line or ':' in line:
                    lines.append(f"    - {line.strip()}")

    return "\n".join(lines)


@app.route("/")
def index():
    return render_template_string(HTML)


@app.route("/feedback", methods=["POST"])
def save_feedback():
    data = request.get_json()
    if not data:
        return jsonify({"error": "No data"}), 400
    equipment = data.get("equipment", "unknown")
    slug = re.sub(r'[^a-z0-9]', '_', equipment.lower()).strip('_')
    fpath = os.path.join(FEEDBACK_DIR, f"{slug}.json")
    try:
        entries = []
        if os.path.exists(fpath):
            with open(fpath) as f:
                entries = json.load(f)
        entries.append(data)
        with open(fpath, "w") as f:
            json.dump(entries, f, indent=2)
        return jsonify({"ok": True, "total_feedback": len(entries)})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/feedback_history/<equipment>")
def feedback_history(equipment):
    slug = re.sub(r'[^a-z0-9]', '_', equipment.lower()).strip('_')
    fpath = os.path.join(FEEDBACK_DIR, f"{slug}.json")
    if not os.path.exists(fpath):
        return jsonify([])
    with open(fpath) as f:
        return jsonify(json.load(f))
    return render_template_string(HTML)


@app.route("/generate", methods=["POST"])
def generate():
    api_key = os.environ.get("ANTHROPIC_API_KEY", "")
    if not api_key:
        return jsonify({"error": "ANTHROPIC_API_KEY not set on server."})

    template_file = request.files.get("template")
    if not template_file or not template_file.filename:
        return jsonify({"error": "No template uploaded."})

    template_bytes = template_file.read()

    pi = {
        "client":       request.form.get("client", ""),
        "project_no":   request.form.get("project_no", ""),
        "project_name": request.form.get("project_name", ""),
        "location":     request.form.get("location", ""),
        "equipment":    request.form.get("equipment", ""),
        "author":       request.form.get("author", ""),
        "vendor1_name": request.form.get("vendor_name_1", "Vendor 1"),
        "vendor2_name": request.form.get("vendor_name_2", "Vendor 2"),
        "vendor3_name": request.form.get("vendor_name_3", "Vendor 3"),
    }

    try:
        wb_tmp = load_workbook(io.BytesIO(template_bytes))
        ws_tmp = wb_tmp.active
        # Warn if template appears to already be filled
        filled_vendor_cells = sum(
            1 for row in ws_tmp.iter_rows(min_row=8, max_row=ws_tmp.max_row)
            for cell in row
            if cell.column in (3,4,5) and cell.value is not None
            and str(cell.value).strip() not in ("","Vendor 1","Vendor 2","Vendor 3","-","By Vendor","N/A")
            and not isinstance(cell, MergedCell)
        )
        if filled_vendor_cells > 30:
            return jsonify({"error": f"Template appears to already be filled ({filled_vendor_cells} vendor cells have data). Please upload a BLANK template — not a previously generated bid tab."})
        row_map = build_row_map(ws_tmp)
        # Load past feedback and inject into system prompt
        equipment_name = request.form.get("equipment", "")
        past_feedback = load_feedback_for_template(equipment_name)
        system_prompt = build_system_prompt(row_map)
        if past_feedback:
            system_prompt = system_prompt + "\n" + past_feedback
    except Exception as e:
        return jsonify({"error": f"Could not read template: {e}"})

    content = ""
    ds = request.files.get("datasheet")
    if ds and ds.filename:
        content += f"=== DATA SHEET ({ds.filename}) ===\n{extract_text(ds)}\n\n"

    # Support additional data sheets (datasheet_2, datasheet_3, etc.)
    ds_count = int(request.form.get("datasheet_count", 1))
    for i in range(2, ds_count + 1):
        dsi = request.files.get(f"datasheet_{i}")
        if dsi and dsi.filename:
            content += f"=== DATA SHEET {i} ({dsi.filename}) ===\n{extract_text(dsi)}\n\n"

    vendor_count = int(request.form.get("vendor_count", 3))
    for i in range(1, vendor_count + 1):
        vf = request.files.get(f"vendor_{i}")
        vn = request.form.get(f"vendor_name_{i}", f"Vendor {i}")
        if vf and vf.filename:
            content += f"=== VENDOR {i} — {vn} ({vf.filename}) ===\n{extract_text(vf)}\n\n"

    if not content.strip():
        return jsonify({"error": "No data sheet or vendor quotes uploaded."})

    try:
        client = anthropic.Anthropic(api_key=api_key)
        resp = client.messages.create(
            model="claude-sonnet-4-5",
            max_tokens=8096,
            system=system_prompt,
            messages=[{"role": "user", "content": content}]
        )
        raw = resp.content[0].text.strip()
        raw = re.sub(r"^```json\s*", "", raw)
        raw = re.sub(r"\s*```$", "", raw)
        data = json.loads(raw)
    except json.JSONDecodeError as e:
        return jsonify({"error": f"AI returned malformed JSON: {e}"})
    except Exception as e:
        return jsonify({"error": f"AI error: {e}"})

    try:
        xlsx_bytes, _ = fill_excel(template_bytes, data, pi)
    except Exception as e:
        return jsonify({"error": f"Excel error: {e}"})

    # Generate versioned filename: BidTabAgent_EQUIPMENT_v1.xlsx, v2.xlsx etc.
    equip = (pi.get("equipment") or "BidTab").strip()
    equip_slug = re.sub(r'[^a-zA-Z0-9]', '_', equip)   # spaces/special → underscore
    equip_slug = re.sub(r'_+', '_', equip_slug).strip('_')  # collapse and trim

    # Find next version number by scanning existing files for this equipment
    existing = [f for f in os.listdir(OUTPUT_DIR)
                if f.startswith(f"BidTabAgent_{equip_slug}_v") and f.endswith(".xlsx")]
    version = len(existing) + 1
    fname = f"BidTabAgent_{equip_slug}_v{version}.xlsx"
    fpath = os.path.join(OUTPUT_DIR, fname)
    with open(fpath, "wb") as f:
        f.write(xlsx_bytes)

    file_count = len([x for x in [ds] + [request.files.get(f"vendor_{i}") for i in range(1, vendor_count+1)] if x and x.filename])
    msg = f"Filled {file_count} files across {len(row_map)} template fields. Yellow cells = not found in vendor quote."
    return jsonify({"filename": fname, "message": msg, "version": version})


@app.route("/download/<fname>")
def download(fname):
    fpath = os.path.join(OUTPUT_DIR, fname)
    if not os.path.exists(fpath):
        return "File not found", 404
    return send_file(fpath, as_attachment=True, download_name=fname,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.errorhandler(500)
def internal_error(e):
    return jsonify({"error": f"Server error: {str(e)}"}), 500

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    print(f"\n  Bid Tab Agent -> http://localhost:{port}\n")
    app.run(host="0.0.0.0", port=port, debug=False)
